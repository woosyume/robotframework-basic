import openpyxl
workbook = openpyxl.load_workbook("/Users/woohyeok.kim/Desktop/study/robotframework-basic/DataDriven/TestData/TestKeyword.xlsx")


def fetch_number_of_rows(sheetname):
    sheet = workbook[sheetname]
    return sheet.max_row


def fetch_cell_data(sheetname, rownum, cellnum):
    sheet = workbook[sheetname]
    cell = sheet.cell(int(rownum), int(cellnum))
    return cell.value


print(fetch_number_of_rows("Sheet1"))
print(fetch_cell_data("Sheet1", 1, 2))