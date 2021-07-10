import openpyxl

# 여기는 load_workbook이 아니네?
workbook = openpyxl.Workbook()
print(workbook.active.title)

sheet = workbook.active
sheet.title = "Hello World"
print(sheet.title)

workbook.create_sheet(title="NewSheet")
newSheet = workbook["NewSheet"]
newSheet['A1'].value = "test value"

sheet['A3'].value = "testuser3"

# Remove sheet
workbook.remove(newSheet)

# Saving
workbook.save("testdata.xlsx")