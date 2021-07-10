import openpyxl

# Load workbook
workbook = openpyxl.load_workbook("testdata.xlsx")
print(workbook.sheetnames)
print("Active sheet= " + workbook.active.title)

sheet = workbook['Sheet2']
print(sheet.title)
print(sheet["A1"].value)

cell = sheet.cell(1, 1)
print(cell.value)
print(cell.row)
print(cell.column)

# Find Rows having data
rows = sheet.max_row
columns = sheet.max_column
print(rows)
print(columns)

# Way1
for i in range(1, rows + 1):
    for j in range(1, columns + 1):
        c = sheet.cell(i, j)
        print(c.value)

# Way2
for row in sheet["A1":"B3"]:
    for cell in row:
        print(cell.value)